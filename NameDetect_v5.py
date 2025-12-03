import os
import io
import re
import asyncio
import aiohttp  # NEW
from dotenv import load_dotenv

import discord
from discord.ext import commands

import easyocr
import cv2
import numpy as np

import difflib
from openpyxl import load_workbook  # For MEMBERLIST.xlsx

# Google Sheets imports
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials

load_dotenv()

# ========= OCR CONFIG =========
LANGS = ["en"]
MIN_CONF = 0.3
SCALE = 2.0
# ==============================

# ========= GOOGLE SHEETS CONFIG =========
# This must match the scope in your token.json
GSHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# From your Sheet URL: https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit#gid=...
GSHEETS_SPREADSHEET_ID = os.getenv("GSHEETS_SPREADSHEET_ID")

# Must match your tab and header in the sheet
GSHEETS_SHEET_NAME = "Data Validation"
GSHEETS_COLUMN_HEADER = "Player IGM"

# Resolve token.json relative to this script file
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GSHEETS_TOKEN_FILE = os.path.join(BASE_DIR, "token.json")
# ========================================

# -------- Intents --------
intents = discord.Intents.default()
intents.message_content = True
intents.members = True  # ✅ needed so we can see guild members

# -------- Bot Setup --------
bot = commands.Bot(command_prefix="!", intents=intents)

print("🔍 Initializing EasyOCR reader...")
ocr_reader = easyocr.Reader(LANGS)
print("✅ EasyOCR reader ready.")


# ========= DOWNLOAD HELPER =========

async def download_attachment_bytes(attachment, retries: int = 3, delay: float = 1.0) -> bytes:
    """
    Robustly download attachment bytes with a few retries to handle
    transient CDN/SSL/ContentLength errors.
    """
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            return await attachment.read()
        except (aiohttp.ClientPayloadError, aiohttp.http_exceptions.ContentLengthError) as e:
            last_exc = e
            print(f"[download_attachment_bytes] Attempt {attempt} failed: {e}")
            if attempt < retries:
                await asyncio.sleep(delay)
            else:
                raise
    raise last_exc


# ========= OCR HELPERS =========

def preprocess_for_ocr_from_bytes(image_bytes: bytes) -> np.ndarray:
    nparr = np.frombuffer(image_bytes, np.uint8)
    img_bgr = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img_bgr is None:
        raise RuntimeError("Failed to decode image bytes with OpenCV.")

    img = cv2.resize(
        img_bgr,
        None,
        fx=SCALE,
        fy=SCALE,
        interpolation=cv2.INTER_CUBIC,
    )
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return gray


def is_probable_name(text: str) -> bool:
    text = text.strip()

    # Basic length filter
    if len(text) < 3:
        return False

    # Must contain at least one letter
    if not any(c.isalpha() for c in text):
        return False

    # Disallow obvious non-name patterns
    if "/" in text:
        return False

    # If there are MANY spaces, it's probably a phrase, not a name
    if text.count(" ") > 1:
        return False

    return True


def normalize_ocr_name(text: str) -> str:
    n = text.strip().strip("[]{}()<>|\"'`.,;:!?")
    n = re.sub(r"^[^A-Za-z0-9]+", "", n)

    if not n:
        return n

    # Remove all whitespace inside (so "Archery202 1" -> "Archery2021")
    n = re.sub(r"\s+", "", n)

    # If it starts with a digit followed by letters somewhere, drop the first digit
    if len(n) >= 2 and n[0].isdigit() and any(c.isalpha() for c in n[1:]):
        n = n[1:]

    # Collapse character spam: "xxxx" -> "xx"
    n = re.sub(r"(.)\1{2,}", r"\1\1", n)

    # Capitalize first letter
    n = n[0].upper() + n[1:]
    return n


def extract_names_from_bytes(image_bytes: bytes):
    preprocessed = preprocess_for_ocr_from_bytes(image_bytes)
    results = ocr_reader.readtext(preprocessed, detail=1)

    print("\n✅ Raw OCR results:")
    for i, (bbox, text, conf) in enumerate(results, start=1):
        print(f"{i}. [{conf:.2f}] {text}")

    names = []
    for bbox, text, conf in results:
        if conf < MIN_CONF:
            continue
        if not is_probable_name(text):
            continue
        cleaned = normalize_ocr_name(text)
        if cleaned:
            names.append(cleaned)

    seen = set()
    unique_names = []
    for n in names:
        key = n.lower()
        if key in seen:
            continue
        seen.add(key)
        unique_names.append(n)

    return unique_names


# ========= MEMBER MATCHING + ROSTER HELPERS =========

def normalize_for_match(s: str) -> str:
    """
    Make string easier to compare:
    - strip spaces
    - lowercase
    - strip leading non-letters (e.g. '4BDNOLEG' -> 'bdnoleg')
    """
    s = s.strip()
    s = re.sub(r"^[^A-Za-z]+", "", s)
    return s.lower()


ROSTER_NAMES = []          # raw names from roster
ROSTER_NORM_MAP = {}       # normalized_name -> original_name


def get_gsheets_creds():
    """
    Load credentials from token.json and refresh if needed.
    Assumes token.json is in the same directory as this script.
    """
    creds = None
    if os.path.exists(GSHEETS_TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(GSHEETS_TOKEN_FILE, GSHEETS_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("[ROSTER] Refreshing Google Sheets access token...")
            creds.refresh(Request())
            # Persist refreshed token back to disk
            with open(GSHEETS_TOKEN_FILE, "w", encoding="utf-8") as f:
                f.write(creds.to_json())
        else:
            raise RuntimeError(
                "[ROSTER] token.json is missing or invalid. "
                "It must contain token, refresh_token, client_id, client_secret, token_uri, and scopes."
            )
    return creds


def load_roster_from_google_sheet():
    """
    Load roster names from a Google Sheet:
    - Spreadsheet: GSHEETS_SPREADSHEET_ID
    - Sheet: GSHEETS_SHEET_NAME
    - Column: GSHEETS_COLUMN_HEADER
    """
    global ROSTER_NAMES, ROSTER_NORM_MAP

    if not GSHEETS_SPREADSHEET_ID:
        print("[ROSTER] GSHEETS_SPREADSHEET_ID not set. Skipping Google Sheets load.")
        return

    try:
        creds = get_gsheets_creds()
    except Exception as e:
        print(f"[ROSTER] Could not get Google Sheets credentials: {e}")
        return

    try:
        service = build("sheets", "v4", credentials=creds)
    except Exception as e:
        print(f"[ROSTER] Failed to initialize Google Sheets service: {e}")
        return

    range_name = f"{GSHEETS_SHEET_NAME}!A:Z"  # wide range; we'll find the column by header
    try:
        sheet = service.spreadsheets()
        result = sheet.values().get(
            spreadsheetId=GSHEETS_SPREADSHEET_ID,
            range=range_name,
        ).execute()
        values = result.get("values", [])
    except Exception as e:
        print(f"[ROSTER] Error reading Google Sheet: {e}")
        return

    if not values:
        print("[ROSTER] No data found in Google Sheet.")
        return

    header_row = values[0]
    try:
        col_index = header_row.index(GSHEETS_COLUMN_HEADER)
    except ValueError:
        print(f"[ROSTER] Column header '{GSHEETS_COLUMN_HEADER}' not found in first row: {header_row}")
        return

    names = []
    for row in values[1:]:
        if len(row) <= col_index:
            continue
        val = row[col_index]
        if val is None:
            continue
        name = str(val).strip()
        if name:
            names.append(name)

    ROSTER_NAMES = names
    ROSTER_NORM_MAP = {normalize_for_match(n): n for n in ROSTER_NAMES}

    print(
        f"[ROSTER] Loaded {len(ROSTER_NAMES)} names from Google Sheet "
        f"(sheet '{GSHEETS_SHEET_NAME}', column '{GSHEETS_COLUMN_HEADER}')."
    )


def load_roster_from_excel(
    path: str = "MEMBERLIST.xlsx",
    sheet_name: str = "Data Validation",
    column_header: str = "Player IGM",
):
    """
    Load roster names from MEMBERLIST.xlsx:
    - Sheet: 'Data Validation'
    - Column: 'Player IGM'
    """
    global ROSTER_NAMES, ROSTER_NORM_MAP

    if not os.path.exists(path):
        print(f"[ROSTER] No Excel file found at {path}. Skipping roster load.")
        return

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[ROSTER] Failed to open {path}: {e}")
        return

    if sheet_name not in wb.sheetnames:
        print(f"[ROSTER] Sheet '{sheet_name}' not found in {path}. Available: {wb.sheetnames}")
        return

    ws = wb[sheet_name]

    # Find the column index for the given header in the first row
    header_row = 1
    col_index = None
    for cell in ws[header_row]:
        if str(cell.value).strip() == column_header:
            col_index = cell.column  # 1-based
            break

    if col_index is None:
        print(f"[ROSTER] Column header '{column_header}' not found in sheet '{sheet_name}'.")
        return

    names = []
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_index)
        val = cell.value
        if val is None:
            continue
        name = str(val).strip()
        if name:
            names.append(name)

    ROSTER_NAMES = names
    ROSTER_NORM_MAP = {normalize_for_match(n): n for n in ROSTER_NAMES}

    print(
        f"[ROSTER] Loaded {len(ROSTER_NAMES)} names from {path} "
        f"(sheet '{sheet_name}', column '{column_header}')."
    )


def correct_with_roster(ocr_name: str, cutoff: float = 0.6) -> str:
    """
    Use the roster to 'snap' an OCR name to the closest known name.
    Returns the corrected roster name if a good match is found,
    otherwise returns the original OCR name.
    """
    if not ROSTER_NORM_MAP:
        # Roster not loaded; no correction
        return ocr_name

    target_norm = normalize_for_match(ocr_name)
    if not target_norm:
        return ocr_name

    candidates = list(ROSTER_NORM_MAP.keys())
    best = difflib.get_close_matches(target_norm, candidates, n=1, cutoff=cutoff)
    if best:
        best_norm = best[0]
        corrected = ROSTER_NORM_MAP[best_norm]
        print(f"[ROSTER] Corrected '{ocr_name}' -> '{corrected}'")
        return corrected

    return ocr_name


async def find_best_member_for_name(guild: discord.Guild, name_for_match: str):
    """
    Simple matcher:
    - Normalize corrected name and member names
    - Try exact normalized match
    - Then substring match
    - Returns first match or None
    """
    if guild is None:
        return None

    target_norm = normalize_for_match(name_for_match)
    if not target_norm:
        return None

    members = []
    async for m in guild.fetch_members(limit=None):
        members.append(m)

    # 1) Exact normalized match on username / display name
    for m in members:
        user_norm = normalize_for_match(m.name)
        display_norm = normalize_for_match(m.display_name)

        if target_norm == user_norm or target_norm == display_norm:
            return m

    # 2) Substring match (e.g. 'bdnoleg' in '4bdnoleg')
    for m in members:
        user_norm = normalize_for_match(m.name)
        display_norm = normalize_for_match(m.display_name)

        if target_norm in user_norm or target_norm in display_norm:
            return m

    return None


# ========= DISCORD BOT =========

@bot.event
async def on_ready():
    print(f"✅ Logged in as {bot.user} (ID: {bot.user.id})")
    print("------")

    # First try Google Sheets
    load_roster_from_google_sheet()

    # Optional fallback to local Excel if Sheets fails / is empty
    if not ROSTER_NAMES:
        print("[ROSTER] Falling back to local Excel roster.")
        load_roster_from_excel(
            path="MEMBERLIST.xlsx",
            sheet_name="Data Validation",
            column_header="Player IGM",
        )


@bot.command()
async def ping(ctx: commands.Context):
    """
    Scan recent messages in this thread/channel,
    collect ALL image attachments, run OCR on each,
    aggregate names, and tag matching members.
    """
    MESSAGE_LIMIT = 50  # how far back to look

    await ctx.send(
        f"🔍 Scanning the last {MESSAGE_LIMIT} messages in this thread for images..."
    )

    image_entries = []  # list of (attachment, message)

    async for msg in ctx.channel.history(limit=MESSAGE_LIMIT):
        if not msg.attachments:
            continue

        for attachment in msg.attachments:
            if (
                attachment.content_type
                and attachment.content_type.startswith("image")
            ) or attachment.filename.lower().endswith(
                (".png", ".jpg", ".jpeg", ".gif", ".webp")
            ):
                image_entries.append((attachment, msg))

    if not image_entries:
        await ctx.send("❌ No image attachments found in the recent messages.")
        return

    await ctx.send(f"✅ Found **{len(image_entries)}** image(s). Running OCR...")

    all_detected_names = []

    # Process each image
    for attachment, msg in image_entries:
        try:
            image_bytes = await download_attachment_bytes(attachment)
        except Exception as e:
            print(f"❌ Failed to download attachment {attachment.filename}: {e}")
            continue

        print(
            f"Processing image '{attachment.filename}' from "
            f"{msg.author} ({len(image_bytes)} bytes)"
        )

        try:
            names = extract_names_from_bytes(image_bytes)
        except Exception as e:
            print(f"❌ OCR error on {attachment.filename}: {e}")
            continue

        if names:
            print(f"➡️ Detected from {attachment.filename}: {names}")
            all_detected_names.extend(names)

    # De-duplicate across all images, preserving order
    if not all_detected_names:
        await ctx.send("😕 I couldn't confidently detect any names from any of the images.")
        return

    unique_names = []
    seen = set()
    for n in all_detected_names:
        key = n.lower()
        if key in seen:
            continue
        seen.add(key)
        unique_names.append(n)

    # ---- MATCH NAMES TO GUILD MEMBERS & TAG THEM (with roster correction) ----
    # Aggregate by member so each user is only tagged once.
    member_hits = {}   # member_id -> {"member": discord.Member, "raw": set(), "corrected": set()}
    unmatched = []     # list of (raw_name, corrected_name)

    for n in unique_names:
        corrected = correct_with_roster(n)
        member = await find_best_member_for_name(ctx.guild, corrected)

        if member:
            entry = member_hits.setdefault(
                member.id,
                {"member": member, "raw": set(), "corrected": set()},
            )
            entry["raw"].add(n)
            entry["corrected"].add(corrected)
        else:
            unmatched.append((n, corrected))

    matched_lines = []

    # Build one line per member (one @mention max)
    for entry in member_hits.values():
        member = entry["member"]
        raw_names = sorted(entry["raw"])
        corrected_names = sorted(entry["corrected"])

        # Usually there’ll be a single corrected name
        if len(corrected_names) == 1:
            corrected = corrected_names[0]
            # Remove duplicates where raw == corrected (case-insensitive)
            raw_unique = [
                r for r in raw_names if r.lower() != corrected.lower()
            ]
            if raw_unique:
                raw_str = ", ".join(f"`{r}`" for r in raw_unique)
                line = f"{raw_str}→`{corrected}`→{member.mention}"
            else:
                line = f"`{corrected}`→{member.mention}"
        else:
            # Edge case: multiple corrected names pointing to same member
            corr_str = ", ".join(f"`{c}`" for c in corrected_names)
            raw_str = ", ".join(f"`{r}`" for r in raw_names)
            line = f"{raw_str}→{corr_str}→{member.mention}"

        matched_lines.append(line)

    # Now build unmatched_lines as before, but from our aggregated list
    unmatched_lines = []
    for raw_name, corrected in unmatched:
        if corrected != raw_name:
            line = f"`{raw_name}`→`{corrected}` (no match)"
        else:
            line = f"`{raw_name}` (no match)"
        unmatched_lines.append(line)

    MAX_DISPLAY = 60  # total lines (matched + unmatched) to keep things short
    display_matched = matched_lines[:MAX_DISPLAY]
    remaining_slots = max(0, MAX_DISPLAY - len(display_matched))
    display_unmatched = unmatched_lines[:remaining_slots]

    extra_matched = len(matched_lines) - len(display_matched)
    extra_unmatched = len(unmatched_lines) - len(display_unmatched)

    response_parts = []
    total_matched_members = len(member_hits)
    response_parts.append(
        f"**Detected {total_matched_members} matched player(s) from {len(image_entries)} image(s).**"
    )

    if display_matched:
        response_parts.append("\n**Matched:**")
        response_parts.extend(display_matched)

    if display_unmatched:
        response_parts.append("\n**Unmatched:**")
        response_parts.extend(display_unmatched)

    if extra_matched > 0 or extra_unmatched > 0:
        response_parts.append(
            f"\n…and {extra_matched} more matched, {extra_unmatched} more unmatched not shown."
        )

    text = "\n".join(response_parts)
    # Safety check: if somehow still too long, hard truncate
    if len(text) > 2000:
        text = text[:1990] + "\n…(truncated)"

    await ctx.send(text)


# -------- Run the bot --------
TOKEN = os.getenv("DISCORD_TOKEN")

if not TOKEN:
    raise RuntimeError("Please set the DISCORD_TOKEN environment variable.")

bot.run(TOKEN)
